from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()
root.title("Python")
root.iconbitmap("computer.ico")
root.geometry("400x400")

# Define Function
def select(a):
    select_label.configure(text=(listbox.get(ANCHOR)))

# Create List
my_list = ["one","two","three"]

# Create Frame
frame = Frame(root)
frame.pack()

# Create scrollbar
scroll = Scrollbar(frame)
scroll.pack(side=RIGHT, fill=Y)

# Create Listbox
listbox = Listbox(frame,width=50,yscrollcommand=scroll.set)
listbox.pack()
scroll.config(command=listbox.yview)

# Add data in listbox
for item in my_list:
    listbox.insert(END,item)

# Create wb & wb
wb = load_workbook("data.xlsx")
ws = wb.active

# Get data from a column of data
col_a = ws["A"]

# Put data from excel to listbox
for item in col_a:
    listbox.insert(END,item.value)

# Put data from excel to listbox
for item in col_a:
    listbox.insert(END,item.value)

# Label
select_label = Label(root,text="Select Something...",font=("Calibri",15))
select_label.pack(pady=10)

# Create listbox binding
listbox.bind("<ButtonRelease-1>",select)

root.mainloop()