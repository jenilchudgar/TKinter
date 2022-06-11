from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()
root.title("Python")
root.iconbitmap("computer.ico")
root.geometry("400x400")

# Create Work Book instance
wb = Workbook()

# Load existing Workbook
wb = load_workbook("data.xlsx")

# Create Active Worksheet
ws = wb.active

# Create Variable for Column A
column_a = ws['A']

# Create Variable for Column B
column_b = ws['B']

def col_a():
    for cell in column_a:
        print(cell.value)

col_a_btn = Button(root,text="Column A",command=col_a)
col_a_btn.pack(pady=10)

def col_b():
    for cell in column_b:
        print(cell.value)

col_b_btn = Button(root,text="Column B",command=col_b)
col_b_btn.pack(pady=10)

# A8 and B8
ws["A8"] = "Jonny"
ws["B8"] = "Strawberry"

wb.save("data.xlsx")

root.mainloop()